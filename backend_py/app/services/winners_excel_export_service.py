"""
Excel export: only winning companies (one row per winner).
Columns: name → tenders won → INN → total sum → phone → region.
"""

from __future__ import annotations

import asyncio
import io
from collections.abc import Awaitable, Callable
from datetime import datetime, timedelta
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.db.session import SessionLocal
from app.models.tender import Tender
from app.models.winner import Winner, CompanyProfile
from app.utils.format_money import format_uzs_total, parse_amount_number
from app.utils.inn import normalize_inn, is_placeholder_company_name
from app.utils.uzex_trade_id import resolve_uzex_trade_id

NO_VALUE = "—"
MAX_TENDER_TITLE = 72
MAX_TENDERS_LISTED = 15

ProgressCallback = Callable[[str], Awaitable[None]]

EXCEL_COLUMNS = [
    ("T/r", "row_num"),
    ("Korxona nomi (g'olib)", "company_name"),
    ("Yutgan tenderlar", "tenders_won"),
    ("Qanday qurilish yoki loyiha yutganligi", "project_titles"),
    ("Tender utkazilgan sana", "tender_dates"),
    ("INN (STIR)", "inn"),
    ("Jami yutgan summa", "amount_total"),
    ("Telefon", "phone"),
    ("Viloyat", "region"),
]


class WinnersExcelExportService:
    """Export only companies that won at least one tender."""

    @staticmethod
    def _company_key(inn: str | None, name: str) -> str:
        if inn and len(inn) == 9:
            return f"inn:{inn}"
        return f"name:{name.strip().lower()[:120]}"

    @staticmethod
    def _deal_title(deal: dict[str, Any], trade_id: int | None) -> str:
        for key in ("trade_name", "product_name", "name", "description"):
            val = deal.get(key)
            if val and str(val).strip():
                return str(val).strip()
        if trade_id:
            return f"Loyiha №{trade_id}"
        return "Tender"

    @staticmethod
    def _shorten_title(title: str) -> str:
        t = " ".join(title.split())
        if len(t) <= MAX_TENDER_TITLE:
            return t
        return t[: MAX_TENDER_TITLE - 1].rstrip() + "…"

    @staticmethod
    def _lot_label(win: dict[str, Any]) -> str:
        lot_no = win.get("display_no") or win.get("trade_id")
        if lot_no:
            return f"№{lot_no}"
        return ""

    @staticmethod
    def _format_tenders_list(wins: list[dict[str, Any]]) -> str:
        lines: list[str] = []
        shown = wins[:MAX_TENDERS_LISTED]
        for i, w in enumerate(shown, 1):
            label = WinnersExcelExportService._lot_label(w)
            title = w.get("title") or "Tender"
            if label:
                lines.append(f"{i}. {label} — {title}")
            else:
                lines.append(f"{i}. {title}")
        rest = len(wins) - len(shown)
        if rest > 0:
            lines.append(f"… va yana {rest} ta tender")
        return "\n".join(lines)

    @staticmethod
    def _format_titles_list(wins: list[dict[str, Any]]) -> str:
        lines: list[str] = []
        shown = wins[:MAX_TENDERS_LISTED]
        for i, w in enumerate(shown, 1):
            title = w.get("title") or "—"
            lines.append(f"{i}. {title}")
        rest = len(wins) - len(shown)
        if rest > 0:
            lines.append(f"… va yana {rest} ta")
        return "\n".join(lines)

    @staticmethod
    def _format_dates_list(wins: list[dict[str, Any]]) -> str:
        lines: list[str] = []
        shown = wins[:MAX_TENDERS_LISTED]
        for i, w in enumerate(shown, 1):
            dt = w.get("deal_date")
            if isinstance(dt, datetime):
                lines.append(f"{i}. {dt.strftime('%d.%m.%Y')}")
            else:
                lines.append(f"{i}. —")
        rest = len(wins) - len(shown)
        if rest > 0:
            lines.append(f"… va yana {rest} ta")
        return "\n".join(lines)

    @staticmethod
    def _deal_region(deal: dict[str, Any]) -> str:
        parts = [deal.get("region_name"), deal.get("district_name")]
        return ", ".join(p for p in parts if p) or ""

    @staticmethod
    def _region_from_trade(trade_data: dict[str, Any]) -> str:
        parts = [
            trade_data.get("delivering_region_name"),
            trade_data.get("delivering_district_name"),
        ]
        return ", ".join(p for p in parts if p) or ""

    @staticmethod
    def _format_phone(raw: str | None) -> str:
        if not raw:
            return NO_VALUE
        p = str(raw).strip()
        if p == NO_VALUE:
            return NO_VALUE
        digits = "".join(c for c in p if c.isdigit())
        if len(digits) == 9:
            return f"+998 {digits}"
        if len(digits) == 12 and digits.startswith("998"):
            return f"+{digits}"
        return p

    @staticmethod
    def _deal_date(deal: dict[str, Any]) -> datetime | None:
        for key in ("deal_date", "sign_date", "end_date", "created_date"):
            raw = deal.get(key)
            if not raw:
                continue
            if isinstance(raw, datetime):
                return raw
            try:
                return datetime.fromisoformat(str(raw).replace("Z", "+00:00").split("+")[0])
            except ValueError:
                continue
        return None

    @staticmethod
    def build_deals_by_company(
        deals_index: dict[int, dict[str, Any]],
    ) -> dict[str, dict[str, Any]]:
        """Group DealsList rows by winner INN (or name)."""
        groups: dict[str, dict[str, Any]] = {}
        for deal in deals_index.values():
            name = (deal.get("provider_name") or "").strip()
            if not name or is_placeholder_company_name(name):
                continue
            inn = normalize_inn(deal.get("provider_inn")) or None
            key = WinnersExcelExportService._company_key(inn, name)
            if key not in groups:
                groups[key] = {
                    "company_name": name,
                    "inn": inn,
                    "phone": None,
                    "wins": {},
                    "regions": set(),
                }
            g = groups[key]
            if inn and not g["inn"]:
                g["inn"] = inn
            if not g["company_name"]:
                g["company_name"] = name

            phone = deal.get("provider_phone") or deal.get("phone")
            if phone and not g["phone"]:
                g["phone"] = str(phone).strip()

            region = WinnersExcelExportService._deal_region(deal)
            if region:
                g["regions"].add(region)

            trade_id = deal.get("trade_id")
            tid = int(trade_id) if trade_id is not None else None
            win_key = str(tid) if tid else f"deal_{id(deal)}"
            amount_raw = deal.get("deal_cost") or deal.get("start_cost")
            g["wins"][win_key] = {
                "trade_id": tid,
                "display_no": deal.get("display_no") or deal.get("trade_no"),
                "title": WinnersExcelExportService._deal_title(deal, tid),
                "amount": parse_amount_number(amount_raw),
                "region": region,
                "deal_date": WinnersExcelExportService._deal_date(deal),
            }
        return groups

    @staticmethod
    def merge_db_winners(
        groups: dict[str, dict[str, Any]],
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> None:
        db = SessionLocal()
        try:
            q = db.query(Winner, Tender).outerjoin(Tender, Tender.id == Winner.tender_id)
            if start_date:
                q = q.filter(Winner.winner_announcement_date >= start_date)
            if end_date:
                q = q.filter(Winner.winner_announcement_date < end_date)

            for w, tender in q.all():
                name = (w.company_name or "").strip()
                if not name or is_placeholder_company_name(name):
                    continue
                inn = normalize_inn(w.company_inn) or None
                key = WinnersExcelExportService._company_key(inn, name)
                if key not in groups:
                    groups[key] = {
                        "company_name": name,
                        "inn": inn,
                        "phone": None,
                        "wins": {},
                        "regions": set(),
                    }
                g = groups[key]
                if w.company_phone and not g["phone"]:
                    g["phone"] = w.company_phone.strip()

                trade_id = None
                if tender:
                    trade_id = resolve_uzex_trade_id(tender.external_id, tender.url)
                win_key = str(trade_id) if trade_id else f"db_{w.id}"
                title = (tender.title if tender else None) or name
                region = (tender.region if tender else "") or ""
                if region:
                    g["regions"].add(region)
                amount = parse_amount_number(w.tender_amount)
                existing = g["wins"].get(win_key)
                if existing:
                    if len(title) > len(existing.get("title", "")):
                        existing["title"] = title
                    if amount and not existing.get("amount"):
                        existing["amount"] = amount
                else:
                    display_no = None
                    if tender and tender.external_id:
                        display_no = str(tender.external_id).strip() or None
                    g["wins"][win_key] = {
                        "trade_id": trade_id,
                        "display_no": display_no,
                        "title": title,
                        "amount": amount,
                        "region": region,
                        "deal_date": w.winner_announcement_date,
                    }
        finally:
            db.close()

    @staticmethod
    def _title_from_trade_data(trade_data: dict[str, Any], trade_id: int) -> str | None:
        """Extract tender title from GetTrade API response."""
        import json as _json

        # Actual tender title (same field as TradeList)
        for key in ("name", "trade_name"):
            val = trade_data.get(key)
            if val and str(val).strip():
                return str(val).strip()

        # Fallback: product name from budget_products
        budget_products_raw = trade_data.get("budget_products")
        if budget_products_raw:
            try:
                products_list = _json.loads(budget_products_raw) if isinstance(budget_products_raw, str) else budget_products_raw
                if isinstance(products_list, list) and products_list:
                    first = products_list[0]
                    name = first.get("Product_Name") or first.get("Description")
                    if name and str(name).strip():
                        return str(name).strip()
            except Exception:
                pass

        for key in ("addon_description", "technical_description", "product_name"):
            val = trade_data.get(key)
            if val and str(val).strip():
                return str(val).strip()

        return None

    @staticmethod
    async def enrich_groups_from_gettrade(
        groups: dict[str, dict[str, Any]],
        api: Any | None = None,
    ) -> None:
        """Fill viloyat and title from GetTrade for each won lot (DealsList has no region/name)."""
        from app.clients.uzex_etender_api import UzexEtenderApiClient

        trade_ids = {
            w["trade_id"]
            for g in groups.values()
            for w in g["wins"].values()
            if w.get("trade_id")
        }
        if not trade_ids:
            return

        close_api = False
        if api is None:
            api = UzexEtenderApiClient()
            close_api = True

        try:
            for i, tid in enumerate(trade_ids, 1):
                try:
                    trade_data = await api.trade_details(tid)
                    region = WinnersExcelExportService._region_from_trade(trade_data)
                    display_no = trade_data.get("display_no")
                    trade_name = WinnersExcelExportService._title_from_trade_data(trade_data, tid)
                    for g in groups.values():
                        for w in g["wins"].values():
                            if w.get("trade_id") != tid:
                                continue
                            if display_no and not w.get("display_no"):
                                w["display_no"] = str(display_no).strip()
                            if region:
                                w["region"] = region
                                g["regions"].add(region)
                            if trade_name and (
                                not w.get("title")
                                or w["title"].startswith("Loyiha №")
                                or w["title"] == "Tender"
                            ):
                                w["title"] = trade_name
                except Exception as e:
                    print(f"⚠️ GetTrade region skip {tid}: {e}")

                if i % 25 == 0:
                    await asyncio.sleep(0.05)
        finally:
            if close_api:
                await api.close()

    @staticmethod
    def _enrich_titles_from_db(groups: dict[str, dict[str, Any]]) -> None:
        db = SessionLocal()
        try:
            for g in groups.values():
                for win in g["wins"].values():
                    tid = win.get("trade_id")
                    if not tid:
                        continue
                    tender = None
                    # Try display_no first (full 14-digit external_id stored in DB)
                    display_no = win.get("display_no")
                    if display_no:
                        tender = (
                            db.query(Tender)
                            .filter(Tender.external_id == str(display_no))
                            .first()
                        )
                    # Fallback: short internal trade_id
                    if not tender:
                        tender = (
                            db.query(Tender)
                            .filter(Tender.external_id == str(tid))
                            .first()
                        )
                    # Fallback: URL contains /lot/{tid}
                    if not tender:
                        tender = (
                            db.query(Tender)
                            .filter(Tender.url.like(f"%/lot/{tid}%"))
                            .first()
                        )
                    if tender and tender.title:
                        win["title"] = tender.title
                    if tender and tender.region:
                        win["region"] = tender.region
                        g["regions"].add(tender.region)
        finally:
            db.close()

    @staticmethod
    def _filter_wins_by_period(
        wins: dict[str, dict[str, Any]],
        start_date: datetime | None,
        end_date: datetime | None,
    ) -> dict[str, dict[str, Any]]:
        if not start_date and not end_date:
            return wins
        filtered: dict[str, dict[str, Any]] = {}
        for k, w in wins.items():
            dt = w.get("deal_date")
            if dt is None:
                filtered[k] = w
                continue
            if start_date and dt < start_date:
                continue
            if end_date and dt >= end_date:
                continue
            filtered[k] = w
        return filtered

    @staticmethod
    def groups_to_rows(
        groups: dict[str, dict[str, Any]],
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for g in groups.values():
            wins_map = WinnersExcelExportService._filter_wins_by_period(
                g["wins"], start_date, end_date
            )
            if not wins_map:
                continue

            wins = sorted(
                wins_map.values(),
                key=lambda x: (x.get("deal_date") or datetime.min),
                reverse=True,
            )
            for w in wins:
                w["title"] = WinnersExcelExportService._shorten_title(w["title"])

            total = sum(w["amount"] or 0 for w in wins)
            regions = sorted(g["regions"]) or sorted(
                {w["region"] for w in wins if w.get("region")}
            )
            region_str = "\n".join(r for r in regions if r) or NO_VALUE

            phone_val = g["phone"] or NO_VALUE
            if phone_val != NO_VALUE:
                phone_val = WinnersExcelExportService._format_phone(phone_val)

            rows.append({
                "company_name": g["company_name"],
                "tenders_won": WinnersExcelExportService._format_tenders_list(wins),
                "project_titles": WinnersExcelExportService._format_titles_list(wins),
                "tender_dates": WinnersExcelExportService._format_dates_list(wins),
                "inn": g["inn"] or NO_VALUE,
                "amount_total": format_uzs_total(total if total > 0 else None),
                "phone": phone_val,
                "region": region_str,
                "_win_count": len(wins),
            })

        rows.sort(key=lambda r: (-r["_win_count"], r["company_name"]))
        return rows

    @staticmethod
    def enrich_phones(rows: list[dict[str, Any]]) -> None:
        db = SessionLocal()
        try:
            inns = {
                r["inn"] for r in rows
                if r.get("inn") and r["inn"] != NO_VALUE
            }
            if not inns:
                return
            norm_inns = {normalize_inn(i) or i for i in inns}
            profiles = {}
            for p in db.query(CompanyProfile).filter(
                CompanyProfile.company_inn.in_(list(norm_inns))
            ).all():
                key = normalize_inn(p.company_inn) if p.company_inn else ""
                if key:
                    profiles[key] = p

            winner_phones: dict[str, str] = {}
            for w in (
                db.query(Winner)
                .filter(Winner.company_inn.in_(list(norm_inns)))
                .filter(Winner.company_phone.isnot(None))
                .order_by(Winner.created_at.desc())
                .all()
            ):
                inn = normalize_inn(w.company_inn) or ""
                if inn and inn not in winner_phones and w.company_phone:
                    winner_phones[inn] = w.company_phone.strip()

            for row in rows:
                if row["phone"] != NO_VALUE:
                    continue
                inn = normalize_inn(row["inn"]) or row["inn"]
                if inn in winner_phones:
                    row["phone"] = WinnersExcelExportService._format_phone(winner_phones[inn])
                    continue
                p = profiles.get(inn)
                if p and p.phone:
                    row["phone"] = WinnersExcelExportService._format_phone(p.phone)
        finally:
            db.close()

    @staticmethod
    async def enrich_phones_via_orginfo(
        rows: list[dict[str, Any]],
        *,
        on_progress: ProgressCallback | None = None,
    ) -> None:
        """Orginfo for every row missing phone (same source as /check_inn)."""
        from app.services.company_enrichment_service import get_company_enrichment_service

        rows_by_inn: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            if row.get("inn") in (None, "", NO_VALUE):
                continue
            inn = normalize_inn(row["inn"]) or row["inn"]
            rows_by_inn.setdefault(inn, []).append(row)

        missing_inns = [
            inn
            for inn, group in rows_by_inn.items()
            if any(r.get("phone") == NO_VALUE for r in group)
        ]
        if not missing_inns:
            return

        service = get_company_enrichment_service()

        async def batch_progress(done: int, total: int) -> None:
            if on_progress:
                await on_progress(f"📞 Orginfo: {done}/{total}")

        enriched = await service.enrich_companies_orginfo_batch(
            missing_inns,
            on_progress=batch_progress,
        )

        for inn, data in enriched.items():
            for row in rows_by_inn.get(inn, []):
                if row.get("phone") == NO_VALUE and data.phone_numbers:
                    row["phone"] = WinnersExcelExportService._format_phone(
                        data.phone_numbers[0]
                    )
                if row.get("region") == NO_VALUE:
                    addr = data.legal_address or data.actual_address or ""
                    region = WinnersExcelExportService._region_from_address(addr)
                    if region:
                        row["region"] = region

    @staticmethod
    def _region_from_address(address: str) -> str:
        """Best-effort viloyat from orginfo address (e.g. 'Toshkent shahri, ...')."""
        addr = " ".join(address.split())
        if not addr:
            return ""
        markers = (
            " viloyati",
            " shahri",
            " tumani",
            " Respublikasi",
        )
        for marker in markers:
            idx = addr.find(marker)
            if idx <= 0:
                continue
            start = addr.rfind(",", 0, idx)
            chunk = addr[start + 1 : idx + len(marker)].strip(" ,")
            if chunk:
                return chunk
        parts = [p.strip() for p in addr.split(",") if p.strip()]
        return parts[0] if parts else ""

    @staticmethod
    def rows_to_excel(rows: list[dict[str, Any]]) -> tuple[bytes, int]:
        if not rows:
            return b"", 0

        export_data = []
        for i, row in enumerate(rows, 1):
            inn = row.get("inn") or NO_VALUE
            if inn != NO_VALUE:
                inn = normalize_inn(inn)
            export_data.append({
                EXCEL_COLUMNS[0][0]: i,
                EXCEL_COLUMNS[1][0]: row["company_name"],
                EXCEL_COLUMNS[2][0]: row["tenders_won"],
                EXCEL_COLUMNS[3][0]: row.get("project_titles", NO_VALUE),
                EXCEL_COLUMNS[4][0]: row.get("tender_dates", NO_VALUE),
                EXCEL_COLUMNS[5][0]: inn,
                EXCEL_COLUMNS[6][0]: row["amount_total"],
                EXCEL_COLUMNS[7][0]: row["phone"],
                EXCEL_COLUMNS[8][0]: row["region"],
            })

        df = pd.DataFrame(export_data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet = "G'oliblar"
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]

            col_map = {name: i + 1 for i, name in enumerate(df.columns)}
            tender_idx = col_map[EXCEL_COLUMNS[2][0]]
            titles_idx = col_map[EXCEL_COLUMNS[3][0]]
            dates_idx = col_map[EXCEL_COLUMNS[4][0]]
            inn_idx = col_map[EXCEL_COLUMNS[5][0]]
            amount_idx = col_map[EXCEL_COLUMNS[6][0]]

            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=inn_idx).number_format = "@"
                for col_idx in (tender_idx, titles_idx, dates_idx):
                    ws.cell(row=r, column=col_idx).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
                ws.cell(row=r, column=amount_idx).alignment = Alignment(
                    vertical="top"
                )
                ws.row_dimensions[r].height = min(
                    15 * max(1, str(ws.cell(row=r, column=tender_idx).value or "").count("\n") + 1),
                    120,
                )

            widths = {
                EXCEL_COLUMNS[0][0]: 5,
                EXCEL_COLUMNS[1][0]: 38,
                EXCEL_COLUMNS[2][0]: 52,
                EXCEL_COLUMNS[3][0]: 52,
                EXCEL_COLUMNS[4][0]: 20,
                EXCEL_COLUMNS[5][0]: 14,
                EXCEL_COLUMNS[6][0]: 28,
                EXCEL_COLUMNS[7][0]: 18,
                EXCEL_COLUMNS[8][0]: 22,
            }
            for title, width in widths.items():
                ws.column_dimensions[get_column_letter(col_map[title])].width = width

        output.seek(0)
        return output.getvalue(), len(export_data)

    @staticmethod
    def _period_bounds(period: str, ref: Optional[datetime] = None) -> tuple[datetime | None, datetime | None, str]:
        ref = ref or datetime.utcnow()
        if period == "daily":
            start = ref.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=1)
            label = start.strftime("%Y-%m-%d")
        elif period == "weekly":
            start = ref - timedelta(days=ref.weekday())
            start = start.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=7)
            label = f"hafta_{start.strftime('%Y%m%d')}"
        elif period == "monthly":
            start = ref.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if ref.month == 12:
                end = ref.replace(year=ref.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                end = ref.replace(month=ref.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
            label = ref.strftime("%Y-%m")
        else:
            return None, None, "barcha"
        return start, end, label

    @classmethod
    async def build_export(
        cls,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        label: str = "goliblar",
        on_progress: ProgressCallback | None = None,
    ) -> tuple[bytes, str, int, int, int]:
        """
        Only winning companies. Data: DealsList + winners DB.
        Returns (excel, filename, company_count, company_count, deals_api_count)
        """
        from app.clients.uzex_etender_api import UzexEtenderApiClient

        async def report(msg: str) -> None:
            if on_progress:
                await on_progress(msg)

        rows: list[dict[str, Any]] = []
        deals_index: dict[int, dict[str, Any]] = {}

        await report("🌐 DealsList yuklanmoqda (g'oliblar)…")
        api = UzexEtenderApiClient()
        try:
            deals_index = await api.build_deals_index(max_pages=50)
            print(f"🌐 DealsList: {len(deals_index)} completed deals")

            await report(f"🌐 {len(deals_index)} bitim\n📦 Bazadan birlashtirish…")

            groups = cls.build_deals_by_company(deals_index)
            cls.merge_db_winners(groups, start_date=start_date, end_date=end_date)
            cls._enrich_titles_from_db(groups)

            await report("📍 Hududlar (GetTrade)…")
            await cls.enrich_groups_from_gettrade(groups, api=api)

            rows = cls.groups_to_rows(groups, start_date=start_date, end_date=end_date)
            cls.enrich_phones(rows)

            await report("📞 Telefonlar (orginfo)…")
            await cls.enrich_phones_via_orginfo(rows, on_progress=on_progress)
        finally:
            await api.close()

        if not rows:
            return b"", f"goliblar_{label}.xlsx", 0, 0, len(deals_index)

        await report(f"🏆 {len(rows)} ta g'olib kompaniya\n📄 Excel…")

        excel, total = cls.rows_to_excel(rows)
        if not excel:
            return b"", f"goliblar_{label}.xlsx", 0, 0, len(deals_index)

        return excel, f"goliblar_{label}.xlsx", total, total, len(deals_index)

    @classmethod
    async def build_period_export(
        cls,
        period: str,
        *,
        all_time: bool = False,
        on_progress: ProgressCallback | None = None,
    ) -> tuple[bytes, str, int, int, int]:
        if all_time:
            start, end, label = None, None, "barcha"
        else:
            start, end, label = cls._period_bounds(period)
        return await cls.build_export(
            start_date=start,
            end_date=end,
            label=label,
            on_progress=on_progress,
        )

    @classmethod
    async def build_all_winners_export(cls, **kwargs) -> tuple[bytes, str, int, int, int]:
        return await cls.build_period_export("monthly", all_time=True, **kwargs)

    @classmethod
    async def build_days_export(cls, days_back: int, **kwargs) -> tuple[bytes, str, int, int, int]:
        start = datetime.utcnow() - timedelta(days=days_back)
        return await cls.build_export(start_date=start, label=f"{days_back}kun", **kwargs)
